#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 同步脚本 — 时间节省数据集中存储

功能:
  1. --init           初始化 Excel 模板（创建表头）
  2. --append <json>  追加单条记录到 Excel
  3. --sync-all       从 records.jsonl 全量同步到 Excel
  4. --read           读取 Excel 全部数据为 JSON（供分析脚本使用）

用法:
  # 初始化 Excel 模板
  python sync_to_excel.py --init --excel "C:/path/to/效贷时间追踪表.xlsx"

  # 追加单条记录
  python sync_to_excel.py --append '{"employee":"吴香康",...}' --excel "C:/path/to/效贷时间追踪表.xlsx"

  # 从 JSONL 全量同步
  python sync_to_excel.py --sync-all --jsonl ~/.workbuddy/data/time-tracking/效贷/records.jsonl --excel "C:/path/to/效贷时间追踪表.xlsx"

  # 读取 Excel 数据为 JSON
  python sync_to_excel.py --read --excel "C:/path/to/效贷时间追踪表.xlsx"

依赖: openpyxl
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：缺少 openpyxl 依赖", file=sys.stderr)
    print("安装命令: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SHEET_NAME = "效贷时间追踪表"

HEADERS = [
    ("记录时间", 22),
    ("日期", 12),
    ("业务线", 8),
    ("员工", 10),
    ("用户故事", 30),
    ("步骤", 12),
    ("步骤代码", 8),
    ("节省小时", 10),
    ("节省人天", 10),
    ("总小时", 10),
    ("备注", 30),
]

HEADER_FILL = PatternFill(start_color="00A870", end_color="00A870", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def init_excel(excel_path: str):
    """初始化 Excel 文件，创建表头"""
    excel_path = os.path.expanduser(excel_path)
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # 写入表头
    for col_idx, (header_name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    print(f"✅ Excel 模板已创建: {excel_path}")
    print(f"   工作表: {SHEET_NAME}")
    print(f"   表头: {', '.join(h for h, _ in HEADERS)}")


def append_record_to_excel(excel_path: str, record: dict):
    """追加单条记录到 Excel"""
    excel_path = os.path.expanduser(excel_path)

    # 如果文件不存在，先初始化
    if not os.path.exists(excel_path):
        init_excel(excel_path)

    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # 找到下一空行
    next_row = ws.max_row + 1
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        next_row = 2  # 只有表头

    # 写入数据
    row_data = [
        record.get("timestamp", ""),
        record.get("date", ""),
        record.get("biz_line", ""),
        record.get("employee", ""),
        record.get("user_story", ""),
        record.get("step", ""),
        record.get("step_code", ""),
        record.get("time_saved_hours", 0),
        record.get("time_saved_pd", 0),
        record.get("total_hours", record.get("time_saved_hours", 0)),
        record.get("remark", ""),
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if col_idx in (8, 9, 10):  # 数值列
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0.00"
        else:
            cell.alignment = Alignment(horizontal="left")

    wb.save(excel_path)


def sync_all_to_excel(jsonl_path: str, excel_path: str):
    """从 JSONL 全量同步到 Excel（覆盖写入）"""
    jsonl_path = os.path.expanduser(jsonl_path)
    excel_path = os.path.expanduser(excel_path)

    if not os.path.exists(jsonl_path):
        print(f"错误: JSONL 文件不存在: {jsonl_path}", file=sys.stderr)
        sys.exit(1)

    # 读取所有记录
    records = []
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue

    if not records:
        print("JSONL 文件中无记录")
        return

    # 初始化 Excel（覆盖）
    init_excel(excel_path)

    # 逐条追加
    for record in records:
        append_record_to_excel(excel_path, record)

    print(f"✅ 已同步 {len(records)} 条记录到 Excel: {excel_path}")


def read_excel_to_json(excel_path: str) -> list:
    """读取 Excel 全部数据为 JSON 列表"""
    excel_path = os.path.expanduser(excel_path)

    if not os.path.exists(excel_path):
        print(f"错误: Excel 文件不存在: {excel_path}", file=sys.stderr)
        return []

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records = []
    for row in rows:
        if not row or not row[0]:
            continue
        record = {
            "timestamp": str(row[0]) if row[0] else "",
            "date": str(row[1]) if len(row) > 1 and row[1] else "",
            "biz_line": str(row[2]) if len(row) > 2 and row[2] else "",
            "employee": str(row[3]) if len(row) > 3 and row[3] else "",
            "user_story": str(row[4]) if len(row) > 4 and row[4] else "",
            "step": str(row[5]) if len(row) > 5 and row[5] else "",
            "step_code": str(row[6]) if len(row) > 6 and row[6] else "",
            "time_saved_hours": float(row[7]) if len(row) > 7 and row[7] else 0,
            "time_saved_pd": float(row[8]) if len(row) > 8 and row[8] else 0,
            "total_hours": float(row[9]) if len(row) > 9 and row[9] else 0,
            "remark": str(row[10]) if len(row) > 10 and row[10] else "",
        }
        records.append(record)

    wb.close()
    return records


def main():
    parser = argparse.ArgumentParser(description="Excel 同步脚本 — 时间节省数据集中存储")

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--init", action="store_true", help="初始化 Excel 模板")
    group.add_argument("--append", type=str, help="追加单条记录（JSON 字符串）")
    group.add_argument("--sync-all", action="store_true", help="从 JSONL 全量同步")
    group.add_argument("--read", action="store_true", help="读取 Excel 为 JSON 输出")

    parser.add_argument("--excel", required=True, help="Excel 文件路径")
    parser.add_argument("--jsonl", default="", help="JSONL 文件路径（--sync-all 时使用）")

    args = parser.parse_args()

    if args.init:
        init_excel(args.excel)

    elif args.append:
        record = json.loads(args.append)
        append_record_to_excel(args.excel, record)
        print(f"✅ 已追加记录到 Excel: {args.excel}")

    elif args.sync_all:
        if not args.jsonl:
            print("错误: --sync-all 需要指定 --jsonl 参数", file=sys.stderr)
            sys.exit(1)
        sync_all_to_excel(args.jsonl, args.excel)

    elif args.read:
        records = read_excel_to_json(args.excel)
        print(json.dumps(records, ensure_ascii=False, indent=2))
        print(f"\n共 {len(records)} 条记录", file=sys.stderr)


if __name__ == "__main__":
    main()
