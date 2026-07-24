"""
测试用例 Excel 生成器（完整版，兼容测试平台导入格式）

根据测试用例 JSON 数据生成符合测试平台导入格式的 Excel 文件。
支持：
1. 前4行固定模板（标题行、说明行、列名行、字段说明行）
2. P0级别用例独立 Sheet
3. 分隔符为空
4. 来源只能是：需求文档、产品用例、补丁用例
5. 从配置文件读取默认值
6. 项目根目录自动定位

使用方法：
    python generate_excel.py <测试用例.json> --caseGroup <功能路径> --version <版本> --manager <责任人>
"""

import os
import sys
import json
import argparse
from datetime import datetime


# ==================== 依赖检查 ====================

def check_dependencies():
    """检查脚本依赖，缺失时给出安装提示"""
    try:
        from openpyxl import Workbook
        return True
    except ImportError:
        print("❌ 缺少依赖: openpyxl")
        print("   安装命令: pip install openpyxl")
        print("   说明: 用于生成 Excel 测试用例文件")
        return False


# 检查依赖
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 导入配置加载器
try:
    from config_loader import get_config, detect_project_root, needs_initialization
except ImportError:
    def get_config(project_root=None):
        return type('Config', (), {
            'get_defaults': lambda: {},
            'validate_required': lambda: ['项目组', '产品名称', '模块路径']
        })()

    def detect_project_root(file_path, max_levels=3):
        return None

    def needs_initialization(project_root):
        return True


# ===== 固定列定义（效贷标准） =====
COLUMNS = [
    ('team', '*项目组'),
    ('caseGroup', '*功能路径（用例分组），层级之间用"-间隔"'),
    ('name', '*功能点（用例名称）'),
    ('preCondition', '功能说明（前置条件）'),
    ('input', '步骤描述，多步骤换行分隔'),
    ('output', '预期结果，多结果换行分隔'),
    ('product', '*产品'),
    ('modulePath', '*模块路径，层级之间用"-间隔"'),
    ('version', '适用版本（产品版本）'),
    ('caseType', '*用例类型'),
    ('source', '来源'),
    ('caseLevel', '用例级别'),
    ('manager', '*责任人，重名的请输入工号'),
    ('autoState', '已实现自动化'),
    ('relateReqCode', '关联用户故事（填写编码），多个用户故事请在单元格内换行输入'),
    ('workload', '工作量（分钟）'),
    ('remarks', '备注'),
    ('separator', '分隔符'),
]


def create_testcase_excel(testcases_data: dict, output_path: str, user_params: dict = None, config=None):
    """
    生成测试用例 Excel 文件（效贷完整格式）

    Args:
        testcases_data: 测试用例 JSON 数据
        output_path: 输出文件路径
        user_params: 用户提供的参数（覆盖默认值）
        config: 配置加载器实例
    """
    if config is None:
        config = get_config()

    defaults = config.get_defaults()

    if user_params is None:
        user_params = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ===== 第1行：标题行（固定） =====
    ws.append(['用例管理 # dmp_testcase'])

    # ===== 第2行：说明行（固定） =====
    ws.append(['1、请将鼠标移到灰色标题行查看字段录入要求\r\n2、红色带星号（*）的字段为必录字段\r\n#SetNULL（启用单元格输入NULL清空字段）'])

    # ===== 第3行：列名行（固定） =====
    ws.append([col[0] for col in COLUMNS])

    # ===== 第4行：字段说明行（固定） =====
    ws.append([col[1] for col in COLUMNS])

    # ===== 第5行起：测试用例数据 =====
    testcases = testcases_data.get('testcases', [])

    p0_testcases = []  # 收集P0级别用例

    for tc in testcases:
        row_data = []

        for col_name, col_label in COLUMNS:
            # 优先使用测试用例中的值
            value = tc.get(col_name)

            # 如果测试用例中没有值，使用用户参数
            if value is None or value == '':
                value = user_params.get(col_name)

            # 如果用户参数也没有，使用配置默认值
            if (value is None or value == '') and col_name in defaults:
                value = defaults.get(col_name, '')

            # ===== 特殊字段处理 =====
            
            # 分隔符：始终为空
            if col_name == 'separator':
                value = ''

            # 来源：只能是需求文档、产品用例、补丁用例
            if col_name == 'source':
                if value not in ['需求文档', '产品用例', '补丁用例']:
                    value = '需求文档'

            # 必填字段仍为空时，使用默认值或提示
            if (value is None or value == '') and col_label.startswith('*'):
                value = defaults.get(col_name, f'请填写{col_name}')

            row_data.append(value)

        ws.append(row_data)

        # 收集P0级别用例
        if tc.get('caseLevel') == 'P0':
            p0_testcases.append(tc)

    # ===== 创建P0级别用例独立Sheet =====
    if p0_testcases:
        ws_p0 = wb.create_sheet('P0级别用例')

        # 前4行固定模板
        ws_p0.append(['用例管理 # dmp_testcase'])
        ws_p0.append(['1、请将鼠标移到灰色标题行查看字段录入要求\r\n2、红色带星号（*）的字段为必录字段\r\n#SetNULL（启用单元格输入NULL清空字段）'])
        ws_p0.append([col[0] for col in COLUMNS])
        ws_p0.append([col[1] for col in COLUMNS])

        # P0用例数据
        for tc in p0_testcases:
            row_data = []
            for col_name, col_label in COLUMNS:
                value = tc.get(col_name)
                if value is None or value == '':
                    value = user_params.get(col_name)
                if (value is None or value == '') and col_name in defaults:
                    value = defaults.get(col_name, '')
                if col_name == 'separator':
                    value = ''
                if col_name == 'source':
                    if value not in ['需求文档', '产品用例', '补丁用例']:
                        value = '需求文档'
                if (value is None or value == '') and col_label.startswith('*'):
                    value = defaults.get(col_name, f'请填写{col_name}')
                row_data.append(value)
            ws_p0.append(row_data)

    # ===== 设置样式 =====
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # ===== 保存文件 =====
    wb.save(output_path)

    # ===== 统计输出 =====
    print(f"Excel文件已生成: {output_path}")
    print(f"   总用例数: {len(testcases)}")
    
    # 按类型统计
    type_stats = {}
    for tc in testcases:
        ct = tc.get('caseType', '功能测试')
        type_stats[ct] = type_stats.get(ct, 0) + 1
    
    for ct, count in type_stats.items():
        print(f"   {ct}: {count}")
    
    if p0_testcases:
        print(f"   P0级别用例: {len(p0_testcases)}（已创建独立Sheet）")


def main():
    # 检查依赖
    if not HAS_OPENPYXL:
        print("\n请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    parser = argparse.ArgumentParser(description='生成测试用例 Excel')

    parser.add_argument('input', help='测试用例 JSON 文件路径')
    parser.add_argument('-o', '--output', help='输出 Excel 文件路径')

    # 用户参数（覆盖默认值）
    parser.add_argument('--caseGroup', required=True, help='功能路径（用例分组）')
    parser.add_argument('--version', required=True, help='适用版本')
    parser.add_argument('--manager', required=True, help='责任人')
    parser.add_argument('--relateReqCode', default='', help='关联用户故事')
    parser.add_argument('--team', default='', help='项目组（覆盖配置）')
    parser.add_argument('--product', default='', help='产品（覆盖配置）')
    parser.add_argument('--modulePath', default='', help='模块路径（覆盖配置）')
    parser.add_argument('--project-root', help='项目根目录（自动定位 .skill 目录）')

    args = parser.parse_args()

    # 自动定位项目根目录
    project_root = args.project_root
    if not project_root:
        project_root = detect_project_root(args.input)

    if project_root:
        print(f"项目根目录: {project_root}")
    else:
        print("未检测到项目根目录，将使用 Skill 默认配置")

    # 检查是否需要初始化
    if project_root and needs_initialization(project_root):
        print("\n" + "=" * 60)
        print("[WARN] 检测到项目未初始化")
        print("=" * 60)
        print(f"项目根目录: {project_root}")
        print(f"缺失配置: {project_root}/.skill/project.yaml")
        print("\n请先初始化项目")
        print("=" * 60)
        return

    # 加载配置
    config = get_config(project_root=project_root)
    missing = config.validate_required()

    # 如果用户没有提供覆盖值且配置缺失，提示用户
    if missing:
        missing_items = []
        if not args.team and '项目组' in missing:
            missing_items.append('项目组 (--team)')
        if not args.product and '产品名称' in missing:
            missing_items.append('产品名称 (--product)')
        if not args.modulePath and '模块路径' in missing:
            missing_items.append('模块路径 (--modulePath)')

        if missing_items:
            print("=" * 60)
            print("[WARN] 配置缺失")
            print("=" * 60)
            print(f"缺失项: {', '.join(missing_items)}")
            if project_root:
                print(f"请在 {project_root}/.skill/project.yaml 中配置")
            else:
                print("请初始化项目或在 .skill/project.yaml 中配置")
            print("或使用命令行参数覆盖")
            print("=" * 60)
            return

    # 读取测试用例 JSON
    with open(args.input, 'r', encoding='utf-8') as f:
        testcases_data = json.load(f)

    # 构建用户参数
    user_params = {
        'caseGroup': args.caseGroup,
        'version': args.version,
        'manager': args.manager,
        'relateReqCode': args.relateReqCode,
        'team': args.team,
        'product': args.product,
        'modulePath': args.modulePath,
    }

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        input_dir = os.path.dirname(args.input)
        input_name = os.path.splitext(os.path.basename(args.input))[0]
        # 去掉 _测试用例 后缀，避免重复
        if input_name.endswith('_测试用例'):
            input_name = input_name[:-len('_测试用例')]
        output_path = os.path.join(input_dir, f'{input_name}_测试用例.xlsx')

    # 生成 Excel
    create_testcase_excel(testcases_data, output_path, user_params, config)


if __name__ == '__main__':
    main()